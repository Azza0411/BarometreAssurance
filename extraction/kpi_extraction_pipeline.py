"""
Pipeline d'extraction des KPI pour chaque document déjà enregistré en base
MySQL (tables `cmf` / `documents`), à partir de quatre tableaux et d'un
paragraphe narratif :
  - le Bilan (voir bilan_kpi_extractor.KPI_DEFINITIONS : Total actif,
    Capitaux propres, Total Passif, et les sections/lignes détaillées du
    Bilan Actif et Passif) ;
  - l'Annexe N°13 "Résultat technique par catégorie d'assurance Non-Vie"
    (voir annexe13_kpi_extractor.KPI_PATTERNS) ;
  - l'Annexe N°12 "Résultat technique par catégorie d'assurance Vie"
    (voir annexe12_kpi_extractor.KPI_PATTERNS) ;
  - les résumés "État de résultat technique" (Vie/Non-Vie) et "État de
    résultat" (résultat global) (voir resultat_kpi_extractor.KPI_PATTERNS) ;
  - le paragraphe "Présentation de la société" (voir
    presentation_kpi_extractor.KPI_NAMES : Date de création, Nombre
    d'actions, Siège social, Effectif).

Pour chaque document :
  - télécharge le PDF en mémoire (aucune écriture du PDF lui-même sur disque) ;
  - extrait tous les KPI des trois tableaux ;
  - enregistre chaque KPI trouvé dans la table MySQL `kpi_values`
    (document_id, tableau, kpi, valeur_nombre ou valeur_texte selon le type
    du KPI) — réexécuter le pipeline met à jour les valeurs existantes sans
    les dupliquer ;
  - pour chaque KPI manquant, consigne l'échec dans un rapport Excel
    data/kpis/echecs_extraction.xlsx (un onglet par KPI, colonnes
    dimensionnées et en-tête figé pour la lisibilité), avec une cause
    best-effort (PDF scanné, document en arabe, structure Takaful, ou motif
    introuvable).
"""

import io
import os
import re
import sys
import time

import pdfplumber
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.repository import (
    ensure_database,
    get_connection,
    get_kpi_values_for_document,
    init_schema,
    list_all_documents,
    save_anomaly,
    save_kpi_value,
)
from extraction.annexe12_kpi_extractor import KPI_PATTERNS as ANNEXE12_KPI_PATTERNS
from extraction.annexe12_kpi_extractor import extract_annexe12_kpis
from extraction.annexe13_kpi_extractor import KPI_PATTERNS as ANNEXE13_KPI_PATTERNS
from extraction.annexe13_kpi_extractor import extract_annexe13_kpis
from extraction.bilan_kpi_extractor import KPI_DEFINITIONS as BILAN_KPI_DEFINITIONS
from extraction.bilan_kpi_extractor import (
    USER_AGENT,
    _find_row_value,
    _is_passif_page,
    _normalizer,
    extract_all_bilan_kpis,
)
from extraction.bvmt_bulletin_kpi_extractor import extract_bulletin_cloture
from extraction.bvmt_kpi_extractor import KPI_NAMES as BVMT_KPI_NAMES
from extraction.bvmt_kpi_extractor import extract_bvmt_kpis
from extraction.cga_kpi_extractor import extract_cga_kpis
from extraction.ftusa_kpi_extractor import KPI_NAMES as FTUSA_KPI_NAMES
from extraction.ftusa_kpi_extractor import extract_ftusa_kpis
from extraction.presentation_kpi_extractor import KPI_NAMES as PRESENTATION_KPI_NAMES
from extraction.presentation_kpi_extractor import extract_presentation_kpis
import extraction.calculated_kpi_extractor as calculated_kpi_extractor
from extraction.data_cleaning import check_yoy_consistency
from extraction.resultat_kpi_extractor import KPI_NAMES as RESULTAT_KPI_NAMES
from extraction.resultat_kpi_extractor import extract_resultat_kpis
from utils.pdf_utils import is_valid_pdf
from utils.pipeline_logging import get_logger, log_json

_logger = get_logger("pipeline")

OUTPUT_DIR = os.path.join("data", "kpis")

# Répertoire racine data/ (ancré sur ce fichier, pas le CWD du process) —
# utilisé pour persister les PDF FTUSA/CGA téléchargés ici (voir
# _run_ftusa/_run_cga). Avant ce correctif, ces PDF n'étaient jamais écrits
# sur disque (contrairement à CMF) : la page Traçabilité KPI ne pouvait donc
# les ouvrir pour aucune de ces deux sources.
_DATA_ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")


def _save_sectoral_pdf(source: str, annee, content: bytes) -> None:
    directory = os.path.join(_DATA_ROOT, source.lower())
    os.makedirs(directory, exist_ok=True)
    with open(os.path.join(directory, f"{source.upper()}_{annee}.pdf"), "wb") as f:
        f.write(content)
FAILURE_REPORT_PATH = os.path.join(OUTPUT_DIR, "echecs_extraction.xlsx")

# Sociétés Takaful : le Bilan y est structuré différemment (colonnes "Fonds
# des Adhérents" / "Entreprise" séparées) et les motifs de recherche actuels
# ne s'y appliquent pas.
TAKAFUL_COMPANIES = {"AL_AMANAH_TAKAFUL", "AT_TAKAFULIA", "ZITOUNA_TAKAFUL"}
ARABIC_RE = re.compile(r"[؀-ۿ]")

BILAN_KPI_NAMES = [definition[0] for definition in BILAN_KPI_DEFINITIONS]
ANNEXE13_KPI_NAMES = list(ANNEXE13_KPI_PATTERNS.keys())
ANNEXE12_KPI_NAMES = list(ANNEXE12_KPI_PATTERNS.keys())
KPI_NAMES = (
    BILAN_KPI_NAMES
    + ANNEXE13_KPI_NAMES
    + ANNEXE12_KPI_NAMES
    + RESULTAT_KPI_NAMES
    + PRESENTATION_KPI_NAMES
)

# Nom du tableau source enregistré avec chaque KPI.
KPI_TABLE_LABEL = {name: "Bilan" for name in BILAN_KPI_NAMES}
KPI_TABLE_LABEL.update(
    {name: "Annexe 13 - Resultat technique Non-Vie" for name in ANNEXE13_KPI_NAMES}
)
KPI_TABLE_LABEL.update(
    {name: "Annexe 12 - Resultat technique Vie" for name in ANNEXE12_KPI_NAMES}
)
KPI_TABLE_LABEL.update(
    {name: "Etat de resultat (technique / global)" for name in RESULTAT_KPI_NAMES}
)
KPI_TABLE_LABEL.update(
    {name: "Presentation de la societe" for name in PRESENTATION_KPI_NAMES}
)


def _get_with_retries(url, timeout, retries=3):
    """Meme approche que les scrapers de collecte (ex: scraping/ftusa_scraper.py) :
    un blip reseau ponctuel ne doit pas faire perdre les KPI d'un document
    pour tout un cycle de pipeline avant la prochaine execution planifiee.

    Verifie aussi que le contenu recu est bien un PDF (`is_valid_pdf`) : un
    lien expire ou une redirection vers une page de login/erreur renvoie
    souvent un statut HTTP 200 avec du HTML au lieu du PDF attendu, ce que
    pdfplumber signalerait plus loin par une exception peu explicite. Tous
    les appels de cette fonction dans ce fichier telechargent un PDF a
    parser -> le controle est fait ici, pas chez l'appelant."""
    for attempt in range(1, retries + 1):
        try:
            response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
            response.raise_for_status()
            if not is_valid_pdf(response.content):
                raise ValueError(f"Contenu recu non-PDF (probablement une page d'erreur) : {url}")
            return response
        except (requests.RequestException, ValueError):
            if attempt == retries:
                raise
            time.sleep(1.5)


# Total actif doit s'equilibrer a la dinar pres avec le total combine
# Capitaux propres + Passif (partie double comptable) : au-dela, l'un des
# deux (ou les deux) est probablement mal extrait.
BALANCE_CHECK_TOLERANCE = 1.0

# Le KPI "Total Passif" (TOTAL_PASSIF_RE dans bilan_kpi_extractor) capture le
# passif SEUL (ex: "Total du Passif" / "Total des passifs"), pas la ligne qui
# s'equilibre reellement avec le Total actif : celle-ci est la ligne
# combinee "Total des capitaux propres et (du) passif(s)" (verifie sur des
# PDF reels COMAR : "TOTAL DES CAPITAUX PROPRES ET DU PASSIF" ; GAT : "Total
# des capitaux propres et passifs") -> KPI dedie pour ce controle, non
# enregistre en base (pas ajoute a KPI_DEFINITIONS/KPI_NAMES) car il ferait
# doublon avec les KPI "Capitaux propres" + "Total Passif" deja exposes.
TOTAL_CP_ET_PASSIF_RE = re.compile(r"^total (des )?capitaux propres et (du )?passifs?\b")


def _check_balance(pdf, kpis):
    """Verifie l'equilibre comptable Total actif = Capitaux propres + Passif.
    Contrairement au garde-fou de plausibilite (bilan_kpi_extractor._is_plausible),
    qui rejette une valeur individuelle implausible, ce controle porte sur la
    coherence entre deux valeurs chacune individuellement plausibles : on ne
    sait pas laquelle des deux est en cause (ou si les deux le sont), donc on
    ne rejette rien ici, on journalise seulement l'ecart pour investigation.
    Renvoie l'ecart absolu si superieur a la tolerance, 0.0 si equilibre,
    None si le controle ne s'applique pas (ligne combinee introuvable)."""
    actif = kpis.get("Total actif")
    if not isinstance(actif, (int, float)):
        return None
    combined_passif = _find_row_value(pdf, TOTAL_CP_ET_PASSIF_RE, header_token=None, page_filter=_is_passif_page)
    if not isinstance(combined_passif, (int, float)):
        return None
    ecart = abs(actif - combined_passif)
    return ecart if ecart > BALANCE_CHECK_TOLERANCE else 0.0


def _extract_all_kpis(pdf):
    kpis = extract_all_bilan_kpis(pdf)
    kpis.update(extract_annexe13_kpis(pdf))
    kpis.update(extract_annexe12_kpis(pdf))
    kpis.update(extract_resultat_kpis(pdf))
    kpis.update(extract_presentation_kpis(pdf))
    return kpis


def _classify_cause(pdf, company_code):
    """Diagnostic best-effort de la cause d'échec, pour orienter un
    traitement manuel ultérieur."""
    if company_code in TAKAFUL_COMPANIES:
        return "Structure Takaful (Bilan combine Fonds des Adherents / Entreprise)"
    sample_text = ""
    for page in pdf.pages[:3]:
        sample_text += page.extract_text() or ""
    if ARABIC_RE.search(sample_text):
        return "Document en arabe"
    total_chars = sum(len(p.chars) for p in pdf.pages[:4])
    if total_chars < 50:
        return "PDF scanne / image (texte non extractible)"
    return "Motif introuvable, ligne absente du document, ou format non reconnu"


def _log_source_result(source, kpi_values_saved, documents_with_kpi, download_errors, **extra):
    """Journalise le resultat d'une sous-pipeline d'extraction dans
    logs/pipeline.log, avec la meme distinction ok/empty que
    pipelines/run_pipeline.py pour la collecte : 0 valeur de KPI enregistree
    n'est pas forcement une panne (aucun nouveau document ce cycle), mais
    doit rester visible separement d'un succes normal plutot que d'etre
    invisible entre deux print() console."""
    event = "kpi_extraction_empty" if kpi_values_saved == 0 else "kpi_extraction_ok"
    log_json(
        _logger, event, source=source, kpi_values_saved=kpi_values_saved,
        documents_with_kpi=documents_with_kpi, download_errors=download_errors, **extra,
    )


def _write_failure_report(failures):
    wb = Workbook()
    wb.remove(wb.active)
    headers = ["Code", "Entreprise", "Annee", "Nom PDF", "Lien", "Cause"]
    col_widths = [10, 45, 8, 30, 70, 55]
    for kpi_name in KPI_NAMES:
        ws = wb.create_sheet(title=kpi_name[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in failures[kpi_name]:
            ws.append(list(row))
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(FAILURE_REPORT_PATH)


def _run_ftusa(conn):
    """Pipeline separe pour les documents de la source FTUSA (sectorielle,
    pas de societe associee) : un seul tableau, l'annexe "Compte
    d'exploitation par branche & par entreprise" (voir
    ftusa_kpi_extractor.extract_ftusa_kpis). Beaucoup plus petit que le
    pipeline CMF (une dizaine de documents au lieu de centaines) : rapporte
    directement sur la console plutot que via un fichier Excel dedie."""
    documents = [doc for doc in list_all_documents(conn) if doc[1] == "FTUSA"]
    print(f"\n===== EXTRACTION KPI FTUSA : {len(documents)} document(s), {len(FTUSA_KPI_NAMES)} KPI =====\n")

    kpi_values_saved = documents_with_kpi = download_errors = 0
    for document_id, _source_nom, _code, _nom_entreprise, nom_pdf, annee, lien in documents:
        print(f"[STEP] FTUSA {annee} : {lien}")
        try:
            response = _get_with_retries(lien, timeout=60)
            _save_sectoral_pdf("FTUSA", annee, response.content)
            with pdfplumber.open(io.BytesIO(response.content)) as pdf:
                kpis = extract_ftusa_kpis(pdf)
        except Exception as exc:
            print(f"  [ERROR] Echec de telechargement/lecture : {exc}")
            download_errors += 1
            continue

        found_count = sum(1 for v in kpis.values() if v is not None)
        for name, value in kpis.items():
            if value is not None:
                save_kpi_value(conn, document_id, "FTUSA - Compte exploitation par branche", name, valeur_nombre=value)
                kpi_values_saved += 1
        if found_count > 0:
            documents_with_kpi += 1
        print(f"  [{'OK' if found_count else 'WARN'}] {found_count}/{len(FTUSA_KPI_NAMES)} KPI enregistres en base")

    print("\n===== RESUME EXTRACTION KPI FTUSA =====")
    print(f"  Documents avec au moins 1 KPI   : {documents_with_kpi}")
    print(f"  Valeurs de KPI enregistrees      : {kpi_values_saved}")
    print(f"  Erreurs telechargement/lecture  : {download_errors}\n")
    _log_source_result("FTUSA", kpi_values_saved, documents_with_kpi, download_errors)

    return {
        "documents_with_kpi": documents_with_kpi,
        "kpi_values_saved": kpi_values_saved,
        "download_errors": download_errors,
    }


def _run_bvmt(conn):
    """Pipeline separe pour les documents de la source BVMT : seuls les
    rapports ESG (PDF, nom_pdf se terminant par ".pdf") necessitent une
    extraction. Les documents de "Status de cotation" n'ont pas de PDF
    associe (nom_pdf sans extension) : leur KPI est deja enregistre pendant
    le scraping (voir scraping.bvmt_scraper.sync_status_cotation)."""
    documents = [
        doc for doc in list_all_documents(conn) if doc[1] == "BVMT" and doc[4].lower().endswith(".pdf")
    ]
    print(f"\n===== EXTRACTION KPI BVMT : {len(documents)} document(s), {len(BVMT_KPI_NAMES)} KPI =====\n")

    kpi_values_saved = documents_with_kpi = download_errors = 0
    for document_id, _source_nom, code, _nom_entreprise, nom_pdf, annee, lien in documents:
        print(f"[STEP] BVMT {code} {annee} : {lien}")
        try:
            response = _get_with_retries(lien, timeout=60)
            with pdfplumber.open(io.BytesIO(response.content)) as pdf:
                kpis = extract_bvmt_kpis(pdf)
        except Exception as exc:
            print(f"  [ERROR] Echec de telechargement/lecture : {exc}")
            download_errors += 1
            continue

        found_count = sum(1 for v in kpis.values() if v is not None)
        for name, value in kpis.items():
            if value is not None:
                save_kpi_value(conn, document_id, "BVMT - Gouvernance", name, valeur_texte=value)
                kpi_values_saved += 1
        if found_count > 0:
            documents_with_kpi += 1
        print(f"  [{'OK' if found_count else 'WARN'}] {found_count}/{len(BVMT_KPI_NAMES)} KPI enregistres en base")

    print("\n===== RESUME EXTRACTION KPI BVMT =====")
    print(f"  Documents avec au moins 1 KPI   : {documents_with_kpi}")
    print(f"  Valeurs de KPI enregistrees      : {kpi_values_saved}")
    print(f"  Erreurs telechargement/lecture  : {download_errors}\n")
    _log_source_result("BVMT", kpi_values_saved, documents_with_kpi, download_errors)

    return {
        "documents_with_kpi": documents_with_kpi,
        "kpi_values_saved": kpi_values_saved,
        "download_errors": download_errors,
    }


def _run_bvmt_bulletin(conn):
    """Pipeline separe pour les bulletins officiels de la cote BVMT
    (documents sectoriels, cmf_id NULL, un par annee, nom_pdf de la forme
    "bulletin_{annee}.pdf" -> voir scraping.bvmt_scraper.sync_market_data) :
    extrait "Cours de l'action - {code}" pour chaque societe cotee reconnue
    (voir bvmt_bulletin_kpi_extractor). Le MNEMO et la denomination BVMT de
    chaque societe, necessaires a cette reconnaissance, sont lus depuis les
    KPI deja enregistres sur son document de profil (cmf_id non NULL, meme
    source) plutot que re-scrapes ici."""
    documents = list_all_documents(conn)
    company_docs = [doc for doc in documents if doc[1] == "BVMT" and doc[2] is not None]

    mnemo_to_code, name_to_code = {}, {}
    for document_id, _source_nom, code, _nom_entreprise, _nom_pdf, _annee, _lien in company_docs:
        kpis = get_kpi_values_for_document(conn, document_id)
        mnemo = kpis.get("Mnemo (BVMT)")
        if isinstance(mnemo, str):
            mnemo_to_code[mnemo] = code
        denomination = kpis.get("Denomination (BVMT)")
        if isinstance(denomination, str):
            name_to_code[_normalizer.clean(denomination)] = code

    bulletin_docs = [doc for doc in documents if doc[1] == "BVMT" and doc[2] is None and doc[4].startswith("bulletin_")]
    print(f"\n===== EXTRACTION KPI BVMT - Bulletins : {len(bulletin_docs)} document(s) =====\n")

    kpi_values_saved = documents_with_kpi = download_errors = 0
    for document_id, _source_nom, _code, _nom_entreprise, _nom_pdf, annee, lien in bulletin_docs:
        print(f"[STEP] BVMT bulletin {annee} : {lien}")
        try:
            response = _get_with_retries(lien, timeout=60)
            with pdfplumber.open(io.BytesIO(response.content)) as pdf:
                cloture_by_code = extract_bulletin_cloture(pdf, mnemo_to_code, name_to_code)
        except Exception as exc:
            print(f"  [ERROR] Echec de telechargement/lecture : {exc}")
            download_errors += 1
            continue

        for code, value in cloture_by_code.items():
            save_kpi_value(
                conn, document_id, "BVMT - Bulletin officiel de la cote", f"Cours de l'action - {code}", valeur_nombre=value
            )
            kpi_values_saved += 1
        if cloture_by_code:
            documents_with_kpi += 1
        print(f"  [{'OK' if cloture_by_code else 'WARN'}] {len(cloture_by_code)} societe(s) trouvee(s)")

    print("\n===== RESUME EXTRACTION KPI BVMT - Bulletins =====")
    print(f"  Documents avec au moins 1 KPI   : {documents_with_kpi}")
    print(f"  Valeurs de KPI enregistrees      : {kpi_values_saved}")
    print(f"  Erreurs telechargement/lecture  : {download_errors}\n")
    _log_source_result("BVMT_BULLETIN", kpi_values_saved, documents_with_kpi, download_errors)

    return {
        "documents_with_kpi": documents_with_kpi,
        "kpi_values_saved": kpi_values_saved,
        "download_errors": download_errors,
    }


def _run_cga(conn):
    """Pipeline separe pour les documents de la source CGA (sectorielle,
    pas de societe associee dans `documents` : les compagnies apparaissent
    en tant que suffixe du nom de KPI, ex: "Nombre d'agences par assureur -
    STAR"). Contrairement aux autres extracteurs, le nombre et les noms des
    KPI varient d'un document a l'autre (une compagnie ou un gouvernorat
    absent d'une annee donnee ne genere simplement pas ce KPI-la) : pas de
    liste KPI_NAMES fixe, donc pas de rapport d'echecs Excel (un KPI
    "manquant" n'est pas necessairement une erreur ici)."""
    documents = [doc for doc in list_all_documents(conn) if doc[1] == "CGA"]
    print(f"\n===== EXTRACTION KPI CGA : {len(documents)} document(s) =====\n")

    kpi_values_saved = documents_with_kpi = download_errors = 0
    for document_id, _source_nom, _code, _nom_entreprise, nom_pdf, annee, lien in documents:
        print(f"[STEP] CGA {annee} : {lien}")
        try:
            response = _get_with_retries(lien, timeout=60)
            _save_sectoral_pdf("CGA", annee, response.content)
            with pdfplumber.open(io.BytesIO(response.content)) as pdf:
                kpis = extract_cga_kpis(pdf)
        except Exception as exc:
            print(f"  [ERROR] Echec de telechargement/lecture : {exc}")
            download_errors += 1
            continue

        found_count = 0
        for name, value in kpis.items():
            if value is not None:
                tableau = "CGA - Annexe 1 - Structure du marche" if name == "Nombre d'assureurs" else "CGA - Annexe 2 - Distribution geographique des agents"
                save_kpi_value(conn, document_id, tableau, name, valeur_nombre=value)
                kpi_values_saved += 1
                found_count += 1
        if found_count > 0:
            documents_with_kpi += 1
        print(f"  [{'OK' if found_count else 'WARN'}] {found_count} KPI enregistres en base")

    print("\n===== RESUME EXTRACTION KPI CGA =====")
    print(f"  Documents avec au moins 1 KPI   : {documents_with_kpi}")
    print(f"  Valeurs de KPI enregistrees      : {kpi_values_saved}")
    print(f"  Erreurs telechargement/lecture  : {download_errors}\n")
    _log_source_result("CGA", kpi_values_saved, documents_with_kpi, download_errors)

    return {
        "documents_with_kpi": documents_with_kpi,
        "kpi_values_saved": kpi_values_saved,
        "download_errors": download_errors,
    }


def run():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ensure_database()
    conn = get_connection()
    init_schema(conn)
    # Les 4 extracteurs ci-dessus ciblent la structure des documents CMF
    # (Bilan, Annexes, États de résultat, Présentation par société) : les
    # documents d'une autre source (ex: FTUSA, sectorielle) ont leur propre
    # pipeline d'extraction (_run_ftusa) et ne passent pas par celui-ci.
    documents = [doc for doc in list_all_documents(conn) if doc[1] == "CMF"]

    print(f"\n===== EXTRACTION KPI : {len(documents)} document(s) en base, {len(KPI_NAMES)} KPI =====\n")

    failures = {name: [] for name in KPI_NAMES}
    kpi_values_saved = documents_with_kpi = download_errors = balance_mismatches = yoy_anomalies = 0

    for document_id, _source_nom, code, nom_entreprise, nom_pdf, annee, lien in documents:
        print(f"[STEP] {code} {annee} : {lien}")
        try:
            response = _get_with_retries(lien, timeout=30)
            with pdfplumber.open(io.BytesIO(response.content)) as pdf:
                kpis = _extract_all_kpis(pdf)
                missing = [name for name in KPI_NAMES if kpis.get(name) is None]
                cause = _classify_cause(pdf, code) if missing else None

                ecart = _check_balance(pdf, kpis)
                if ecart:
                    print(f"  [WARN] Desequilibre Bilan : Total actif != Capitaux propres + Passif (ecart={ecart:,.3f} TND)")
                    details = {"total_actif": kpis.get("Total actif"), "ecart": round(ecart, 3)}
                    log_json(_logger, "balance_check_failed", company=code, annee=annee, **details)
                    # Persiste en base (pas seulement dans logs/pipeline.log) pour que
                    # les pages Qualite/Anomalies puissent lire ce signal (voir
                    # api/services/quality.py, pipeline_audit.py, anomalies_service.py).
                    save_anomaly(
                        conn, source="extraction_balance", gravite="erreur",
                        code=code, annee=annee, kpi="Total actif", details=details,
                    )
                    balance_mismatches += 1

                for flag in check_yoy_consistency(conn, code, annee, kpis):
                    print(
                        f"  [WARN] Variation YoY suspecte {flag['kpi']} : "
                        f"{flag['valeur_precedente']:,.0f} ({flag['annee_precedente']}) -> "
                        f"{flag['valeur_actuelle']:,.0f} ({annee}) [{flag['variation_pct']:+.1f}%]"
                    )
                    log_json(_logger, "yoy_anomaly", company=code, annee=annee, **flag)
                    save_anomaly(
                        conn, source="extraction_yoy", gravite="avertissement",
                        code=code, annee=annee, kpi=flag["kpi"], details=flag,
                    )
                    yoy_anomalies += 1
        except Exception as exc:
            print(f"  [ERROR] Echec de telechargement/lecture : {exc}")
            download_errors += 1
            for name in KPI_NAMES:
                failures[name].append((code, nom_entreprise, annee, nom_pdf, lien, f"Erreur : {exc}"))
            continue

        for name in missing:
            failures[name].append((code, nom_entreprise, annee, nom_pdf, lien, cause))

        found_count = len(KPI_NAMES) - len(missing)
        if found_count > 0:
            for name in KPI_NAMES:
                value = kpis.get(name)
                if value is not None:
                    if isinstance(value, str):
                        save_kpi_value(conn, document_id, KPI_TABLE_LABEL[name], name, valeur_texte=value)
                    else:
                        save_kpi_value(conn, document_id, KPI_TABLE_LABEL[name], name, valeur_nombre=value)
                    kpi_values_saved += 1
            documents_with_kpi += 1
            print(f"  [OK] {found_count}/{len(KPI_NAMES)} KPI enregistres en base")
        else:
            print(f"  [WARN] Aucun KPI trouve ({cause})")

    _write_failure_report(failures)

    print("\n===== RESUME EXTRACTION KPI =====")
    print(f"  Documents avec au moins 1 KPI   : {documents_with_kpi}")
    print(f"  Valeurs de KPI enregistrees      : {kpi_values_saved}")
    print(f"  Erreurs telechargement/lecture  : {download_errors}")
    print(f"  Desequilibres Bilan (Actif!=Passif) : {balance_mismatches}")
    print(f"  Anomalies annee sur annee (YoY)  : {yoy_anomalies}")
    for name in KPI_NAMES:
        print(f"  Echecs {name:60s} : {len(failures[name])}")
    print(f"  Rapport d'echecs                : {FAILURE_REPORT_PATH}\n")
    _log_source_result(
        "CMF", kpi_values_saved, documents_with_kpi, download_errors,
        nb_documents=len(documents), nb_failures_total=sum(len(v) for v in failures.values()),
        balance_mismatches=balance_mismatches, yoy_anomalies=yoy_anomalies,
    )

    ftusa_stats = _run_ftusa(conn)
    bvmt_stats = _run_bvmt(conn)
    bvmt_bulletin_stats = _run_bvmt_bulletin(conn)
    cga_stats = _run_cga(conn)

    # Modelisation : KPI calcules (ratios, ROE/ROA, part de marche, taux de
    # penetration...) a partir des KPI bruts que les 5 etapes ci-dessus
    # viennent d'enregistrer. Doit s'executer APRES elles (CMF, FTUSA, BVMT,
    # BVMT bulletin, CGA sont tous des entrees possibles selon la famille de
    # calcul) - jusqu'ici jamais appele automatiquement, seulement via
    # `python -m extraction.calculated_kpi_extractor` a la main.
    print("\n===== MODELISATION : KPI CALCULES =====\n")
    calculated_stats = calculated_kpi_extractor.run(conn)
    calculated_total_saved = sum(
        stats.get("kpi_values_saved", 0) for stats in calculated_stats.values()
    )
    log_json(_logger, "modelisation_done", **{
        family: stats.get("kpi_values_saved", 0) for family, stats in calculated_stats.items()
    })

    conn.close()

    return {
        "documents_with_kpi": documents_with_kpi,
        "kpi_values_saved": kpi_values_saved,
        "download_errors": download_errors,
        "balance_mismatches": balance_mismatches,
        "yoy_anomalies": yoy_anomalies,
        "failures": failures,
        "ftusa": ftusa_stats,
        "bvmt": bvmt_stats,
        "bvmt_bulletin": bvmt_bulletin_stats,
        "cga": cga_stats,
        "calculated": calculated_stats,
        "calculated_kpi_values_saved": calculated_total_saved,
    }


if __name__ == "__main__":
    run()
