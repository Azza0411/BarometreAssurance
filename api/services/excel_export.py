"""Génération de l'export Excel — Analyse Comparative — avec formules
recalculables pour les ratios simples (ROE, ROA, Ratio de frais, PDM).

Ratio combiné et Ratio de sinistralité restent des valeurs figées : leur
calcul réel (kpi_builder.py::build_company_row) enchaîne ~15 règles de
repli (agrégation Vie/Non-Vie, détection de valeurs non fiables,
invalidation croisée RC/RF...) — les reproduire comme formules tableur
risquerait de produire des chiffres qui divergent silencieusement de
l'application, la classe de bug precisement traquee dans Qualité Data
cette session."""

import io
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from api.utils.formatters import kpis_by_year
from database.repository import get_connection, list_documents_by_source, get_kpi_values_for_document

DARK = "2E2E38"
YELLOW = "FFE600"
LIGHT = "F4F4F7"

# Même copie locale que api/services/pdf_export.py (voir son commentaire :
# Dockerfile.api ne copie pas frontend/, donc pas d'accès à frontend/public/) —
# "EyForDarkBg.png", pas "EyDark.png" : ce dernier est un WebP renommé .png,
# opaque (aucune transparence — vérifié via PIL) avec un fond BLANC PLEIN,
# invisible en PDF seulement par chance (le mask="auto" de ReportLab
# masquait le défaut) mais collé tel quel par openpyxl — un bloc blanc plein
# apparaissait sur le bandeau sombre. Voir le commentaire détaillé dans
# pdf_export.py pour l'origine exacte du fichier de remplacement.
_LOGO_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "assets", "EyForDarkBg.png",
)

# Lignes réservées au bandeau de titre en haut de feuille (logo + titre +
# sous-titre + espacement), ajoutées le 2026-08-19 — jusque-là le fichier
# démarrait directement sur la ligne d'en-tête du tableau, sans aucune
# identification visuelle du rapport (contrairement aux exports PDF, qui ont
# toujours eu ce bandeau). Toutes les lignes du tableau (en-tête + données)
# sont décalées de cette valeur.
_TITLE_ROWS = 3
_HEADER_ROW = _TITLE_ROWS + 1
_DATA_START_ROW = _HEADER_ROW + 1


def _logo_as_png():
    """Ré-encodage défensif en PNG véritable avant de passer l'image à
    openpyxl : `EyForDarkBg.png` est déjà un vrai PNG (contrairement à
    l'ancien `EyDark.png`, un WebP renommé .png qui faisait échouer
    l'insertion — "KeyError: '.webp'" côté manifeste du xlsx, openpyxl
    déterminant le type MIME depuis l'EXTENSION du chemin, pas le contenu
    réel), mais on garde ce passage pour rester robuste si le fichier
    source change à nouveau."""
    if not os.path.exists(_LOGO_PATH):
        return None
    try:
        buf = io.BytesIO()
        PILImage.open(_LOGO_PATH).convert("RGBA").save(buf, format="PNG")
        buf.seek(0)
        return PILImage.open(buf)
    except Exception:
        return None


def _write_title_banner(ws, last_col, title, subtitle):
    last_col_letter = get_column_letter(last_col)
    banner_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    for r in (1, 2):
        for col_idx in range(1, last_col + 1):
            ws.cell(row=r, column=col_idx).fill = banner_fill
    ws.merge_cells(f"B1:{last_col_letter}1")
    ws.merge_cells(f"B2:{last_col_letter}2")
    ws["B1"] = title
    ws["B1"].font = Font(color="FFFFFF", bold=True, size=15, name="Calibri")
    ws["B1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B2"] = subtitle
    ws["B2"].font = Font(color=YELLOW, size=10, name="Calibri")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6
    pil_logo = _logo_as_png()
    if pil_logo is not None:
        try:
            img = XLImage(pil_logo)
            img.height = 58  # agrandi le 2026-08-19 (retour utilisateur)
            img.width = 58 * 0.864  # ratio largeur/hauteur du PNG source (768x889)
            ws.add_image(img, "A1")
        except Exception:
            pass
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DARK

_RAW_KPIS = [
    "Primes émises par assurance", "Résultat Net", "Capitaux propres",
    "Total actif", "Charges d'acquisition et de gestion nettes",
]


def _header_fill():
    return PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")


def _header_font():
    return Font(color=YELLOW, bold=True, name="Calibri", size=10)


def _thin_border():
    side = Side(style="thin", color="DDDDE3")
    return Border(left=side, right=side, top=side, bottom=side)


def build_analyse_comparative_xlsx(annee):
    conn = get_connection()
    try:
        raw_by_company = {}
        for doc_id, _cmf_id, code, doc_annee in list_documents_by_source(conn, "CMF"):
            if doc_annee != annee or not code:
                continue
            kpis = get_kpi_values_for_document(conn, doc_id)
            if any(kpis.get(k) is not None for k in _RAW_KPIS):
                raw_by_company[code] = kpis

        ftusa = kpis_by_year(conn, "FTUSA").get(annee, {})
        total_marche = ftusa.get("Total Primes émises")

        from api.app import app
        from api.routes.comparative import analyse_comparative
        with app.test_request_context(f"/x?annee={annee}"):
            comp_data = analyse_comparative().get_json()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Analyse Comparative {annee}"

    headers = [
        "Compagnie", "Primes émises (MDT)", "Résultat net (MDT)",
        "Capitaux propres (MDT)", "Total actif (MDT)",
        "Charges d'acquisition et de gestion nettes (MDT)",
        "PDM (%)", "ROE (%)", "ROA (%)", "Ratio de frais de gestion (%)",
        "Ratio combiné (%)*", "Ratio de sinistralité (%)*",
    ]
    # "Total primes marché" est une constante du MARCHÉ (une seule valeur,
    # pas une par compagnie) référencée par les formules PDM ci-dessous —
    # jusqu'ici affichée dans une colonne isolée, sans bordure ni fond,
    # flottant à droite du tableau plutôt que d'en faire visiblement partie
    # (signalé par l'utilisateur le 2026-08-19 : "ça doit appartenir au
    # même tableau, ce qui n'est pas le cas"). Traitée maintenant comme une
    # 13e colonne à part entière : même style d'en-tête (fond sombre/texte
    # jaune) que les autres colonnes, cellule de donnée fusionnée sur toute
    # la hauteur du tableau (une seule valeur, pas une par ligne) avec la
    # même bordure/police que les cellules de données.
    total_marche_col = len(headers) + 1
    all_headers = headers + ["Total primes marché (MDT)"]

    _write_title_banner(
        ws, len(all_headers), f"Analyse Comparative — Exercice {annee}",
        f"FS Market Intelligence · Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
    )
    for col_idx, value in enumerate(all_headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col_idx, value=value)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[_HEADER_ROW].height = 32

    total_marche_ref = f"${get_column_letter(total_marche_col)}${_DATA_START_ROW}"

    codes_sorted = sorted(
        comp_data.keys(),
        key=lambda c: (comp_data[c].get("pdm") if comp_data[c].get("pdm") is not None else -1),
        reverse=True,
    )

    row_idx = _DATA_START_ROW
    for code in codes_sorted:
        kpis = raw_by_company.get(code, {})
        comp = comp_data[code]

        def mdt(kpi_name):
            v = kpis.get(kpi_name)
            return round(v / 1_000_000, 3) if v is not None else None

        primes = mdt("Primes émises par assurance")
        resultat = mdt("Résultat Net")
        capitaux = mdt("Capitaux propres")
        actif = mdt("Total actif")
        charges_frais = mdt("Charges d'acquisition et de gestion nettes")

        c_primes    = f"B{row_idx}"
        c_resultat  = f"C{row_idx}"
        c_capitaux  = f"D{row_idx}"
        c_actif     = f"E{row_idx}"
        c_charges   = f"F{row_idx}"

        row_values = [
            code, primes, resultat, capitaux, actif, charges_frais,
            f"=IF(OR({c_primes}=\"\",{total_marche_ref}=\"\",{total_marche_ref}=0),\"\",{c_primes}/{total_marche_ref}*100)",
            f"=IF(OR({c_resultat}=\"\",{c_capitaux}=\"\",{c_capitaux}=0),\"\",{c_resultat}/{c_capitaux}*100)",
            f"=IF(OR({c_resultat}=\"\",{c_actif}=\"\",{c_actif}=0),\"\",{c_resultat}/{c_actif}*100)",
            f"=IF(OR({c_charges}=\"\",{c_primes}=\"\",{c_primes}=0),\"\",ABS({c_charges})/{c_primes}*100)",
            comp.get("ratio_combine"),
            comp.get("ratio_sp"),
        ]

        fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid") if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _thin_border()
            cell.font = Font(name="Calibri", size=10, bold=(col_idx == 1), color=DARK)
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
            if fill:
                cell.fill = fill
            if col_idx > 1:
                cell.number_format = "0.0"
        row_idx += 1
    last_data_row = row_idx - 1

    # Cellule "Total primes marché" : une seule valeur pour tout le marché,
    # fusionnée sur toute la hauteur des données (pas une par compagnie) —
    # même bordure/police que les autres cellules de données pour bien
    # lire comme la 13e colonne du même tableau, pas un fragment à part.
    if last_data_row >= _DATA_START_ROW:
        ws.merge_cells(start_row=_DATA_START_ROW, start_column=total_marche_col,
                        end_row=last_data_row, end_column=total_marche_col)
    total_marche_cell = ws.cell(row=_DATA_START_ROW, column=total_marche_col,
                                 value=round(total_marche / 1_000_000, 3) if total_marche else None)
    total_marche_cell.font = Font(name="Calibri", size=11, bold=True, color=DARK)
    total_marche_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_marche_cell.number_format = "0.0"
    total_marche_cell.fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
    for r in range(_DATA_START_ROW, last_data_row + 1):
        ws.cell(row=r, column=total_marche_col).border = _thin_border()

    note = ws.cell(row=_HEADER_ROW, column=11, value="Ratio combiné (%)*")
    note.comment = Comment(
        "Valeur figée, calculée côté serveur — identique à celle affichée dans "
        "l'application. Non recalculable ici : sa formule réelle enchaîne "
        "plusieurs règles de repli (agrégation Vie/Non-Vie, détection de "
        "valeurs non fiables) trop complexes pour une formule tableur fiable.",
        "FS Market Intelligence", height=140, width=260,
    )
    note2 = ws.cell(row=_HEADER_ROW, column=12, value="Ratio de sinistralité (%)*")
    note2.comment = Comment(
        "Valeur figée — même raison que Ratio combiné (voir commentaire de cette colonne).",
        "FS Market Intelligence", height=80, width=260,
    )
    ws.cell(row=_HEADER_ROW, column=6).comment = Comment(
        "Convention comptable : les charges sont stockées en négatif (sortie "
        "de trésorerie). Le Ratio de frais applique ABS() sur cette colonne.",
        "FS Market Intelligence", height=90, width=240,
    )

    for col_idx, width in enumerate([14, 16, 16, 18, 16, 30, 10, 10, 10, 16, 16, 18, 20], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{_DATA_START_ROW}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
